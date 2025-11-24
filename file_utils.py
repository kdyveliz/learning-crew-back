# file_utils.py
import os
import logging
from typing import Optional

# Excel 파일 읽기 라이브러리 임포트 시도
try:
    import pandas as pd

    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    try:
        import openpyxl

        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


def clean_filename(filename: str) -> str:
    """파일명에서 (1), (2) 같은 번호 패턴을 제거합니다.

    예: "광주_4반_송주헌 (1).xlsx" -> "광주_4반_송주헌.xlsx"
    """
    import re

    # 확장자 분리
    name, ext = os.path.splitext(filename)

    # (숫자) 패턴 제거 - 앞뒤 공백도 함께 제거
    cleaned_name = re.sub(r"\s*\(\d+\)\s*", "", name).strip()

    # 정제된 이름 + 확장자 반환
    return cleaned_name + ext


def read_file_with_encoding(file_path: str) -> str:
    """파일을 읽어서 반환 (여러 인코딩 시도)"""
    encodings = ["utf-8", "utf-8-sig", "cp949", "euc-kr", "latin1"]

    for encoding in encodings:
        try:
            with open(file_path, "r", encoding=encoding) as f:
                return f.read()
        except (UnicodeDecodeError, FileNotFoundError):
            continue

    logger.error(f"파일을 찾을 수 없거나 인코딩을 확인할 수 없습니다: {file_path}")
    raise FileNotFoundError(
        f"파일을 찾을 수 없거나 인코딩을 확인할 수 없습니다: {file_path}"
    )


def read_excel_file(file_path: str) -> str:
    """Excel 파일을 읽어서 텍스트로 반환"""
    if HAS_PANDAS:
        try:
            excel_file = pd.ExcelFile(file_path)
            content_parts = []

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                content_parts.append(f"[시트: {sheet_name}]\n")
                content_parts.append(df.to_string(index=False))
                content_parts.append("\n\n")

            return "\n".join(content_parts)
        except Exception as e:
            raise Exception(f"Pandas로 Excel 파일을 읽는 중 오류 발생: {e}")

    elif HAS_OPENPYXL:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)
            content_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                content_parts.append(f"[시트: {sheet_name}]\n")

                for row in ws.iter_rows(values_only=True):
                    row_str = "\t".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    content_parts.append(row_str)
                content_parts.append("\n\n")

            return "\n".join(content_parts)
        except Exception as e:
            raise Exception(f"Openpyxl로 Excel 파일을 읽는 중 오류 발생: {e}")
    else:
        raise ImportError(
            "Excel 파일을 읽으려면 pandas 또는 openpyxl이 필요합니다. 'pip install pandas openpyxl'을 실행하세요."
        )


def load_system_prompt(file_path: str) -> str:
    """시스템 프롬프트 파일을 읽고, UTF-8로 다시 저장 (원본 기능 유지)"""
    content = read_file_with_encoding(file_path)  # FileNotFoundError 여기서 발생 가능
    try:
        # UTF-8로 다시 저장하여 다음번에는 UTF-8로 읽을 수 있도록 함
        with open(file_path, "w", encoding="utf-8") as f_out:
            f_out.write(content)
    except Exception as e:
        logger.warning(f"시스템 프롬프트를 UTF-8로 다시 저장하는 데 실패했습니다: {e}")
    return content


def find_file_by_keywords(search_path: str, keywords: list[str]) -> Optional[str]:
    """지정된 경로에서 키워드가 모두 포함된 파일명을 찾아 전체 경로를 반환합니다.

    파일명에서 (1), (2) 등의 번호 패턴을 무시하고 검색합니다.
    예: "계획서(1).xlsx", "결과보고서(2).txt" 등도 매칭됩니다.
    """
    import re

    if not os.path.exists(search_path):
        logger.warning(f"검색 경로를 찾을 수 없습니다: {search_path}")
        return None

    for file in os.listdir(search_path):
        # 확장자 분리
        file_name, file_ext = os.path.splitext(file)

        # 파일명에서 (숫자) 패턴 제거 (예: "계획서(1)" -> "계획서")
        cleaned_name = re.sub(r"\(\d+\)", "", file_name)

        # 정제된 파일명 + 확장자로 재구성
        cleaned_file = cleaned_name + file_ext

        # 키워드 매칭 (원본 파일명과 정제된 파일명 모두 확인)
        if all(keyword in file for keyword in keywords) or all(
            keyword in cleaned_file for keyword in keywords
        ):
            logger.info(f"파일 발견: {file}")
            return os.path.join(search_path, file)

    logger.warning(
        f"키워드 {keywords}에 해당하는 파일을 {search_path}에서 찾지 못했습니다."
    )
    return None


def get_file_content(file_path: str) -> str:
    """파일 확장자에 따라 적절한 읽기 함수를 호출하여 내용을 반환합니다."""
    file_ext = os.path.splitext(file_path)[1].lower()

    if file_ext in [".xlsx", ".xls"]:
        logger.info(f"Excel 파일 읽기 시도: {file_path}")
        return read_excel_file(file_path)
    else:
        logger.info(f"텍스트 파일 읽기 시도: {file_path}")
        return read_file_with_encoding(file_path)


# --- analyzer_logic.py에서 이동된 로직 ---

import io
import zipfile
from PIL import Image
from typing import Optional
from fastapi import UploadFile


def extract_images_from_excel(file_bytes: bytes) -> list[Image.Image]:
    """Excel 파일(ZIP 구조)에서 이미지를 추출합니다."""
    images = []
    MIN_IMAGE_SIZE = 15000  # 15KB 미만 무시

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            media_files = [
                f
                for f in z.namelist()
                if f.startswith("xl/media/") and not f.endswith("/")
            ]

            for file_name in media_files:
                try:
                    img_data = z.read(file_name)
                    if len(img_data) < MIN_IMAGE_SIZE:
                        continue
                    img = Image.open(io.BytesIO(img_data))
                    images.append(img)
                except Exception:
                    continue

            return images
    except Exception:
        return []


async def read_upload_file_content(file: UploadFile) -> str:
    """FastAPI UploadFile 객체에서 텍스트 내용을 읽어옵니다."""
    await file.seek(0)
    content_bytes = await file.read()
    filename_lower = file.filename.lower()

    if filename_lower.endswith(".xlsx"):
        file_stream = io.BytesIO(content_bytes)
        # pandas 의존성 확인
        if HAS_PANDAS:
            try:
                excel_data = pd.read_excel(
                    file_stream, sheet_name=None, engine="openpyxl"
                )
                report_content = ""

                for sheet_name, df in excel_data.items():
                    # 1. 데이터 정제
                    df = df.fillna("")
                    # 2. 헤더 정제
                    new_columns = [
                        "" if "Unnamed" in str(col) else str(col) for col in df.columns
                    ]
                    df.columns = new_columns

                    # 3. 문자열 변환
                    try:
                        table_text = df.to_string(index=False)
                    except:
                        table_text = str(df)

                    report_content += f"\n### 시트명: {sheet_name}\n{table_text}\n"

                return report_content
            except Exception as e:
                logger.error(f"Pandas 읽기 실패: {e}")
                # Fallback to openpyxl if needed or re-raise
                pass

        # Fallback or if no pandas (using openpyxl directly if implemented,
        # but for now reusing the logic from analyzer_logic which used pandas)
        # If pandas is missing, this part might fail if we don't have alternative logic here.
        # Assuming HAS_PANDAS is true for the environment as per original code.
        return "Excel file content (Pandas required)"

    elif filename_lower.endswith(".txt") or filename_lower.endswith(".csv"):
        try:
            return content_bytes.decode("utf-8")
        except:
            return content_bytes.decode("cp949", errors="ignore")
    else:
        return content_bytes.decode("utf-8", errors="ignore")


def extract_attendance_list_from_excel(file_path: str) -> list[str]:
    """Excel 파일에서 '스터디 참석자' 또는 '참석명단' 정보를 추출합니다.

    1. 컬럼 헤더에 키워드가 있는 경우 (테이블 형식)
    2. 셀 데이터에 키워드가 있는 경우 (폼 형식, 바로 오른쪽 셀 데이터 추출)

    Args:
        file_path: Excel 파일 경로

    Returns:
        참석자 이름 리스트 (없으면 빈 리스트)
    """
    KEYWORDS = ["스터디 참석자", "참석명단", "스터디장"]

    if HAS_PANDAS:
        try:
            # Excel 파일 로드
            xls = pd.ExcelFile(file_path)

            all_attendees = []

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(
                    xls, sheet_name=sheet_name, header=None
                )  # 헤더 없이 읽어서 전체 탐색

                # 로그에 엑셀 내용 출력 (디버깅용)
                logger.info(f"--- [Sheet: {sheet_name}] Content Preview ---")
                logger.info(f"\n{df.head(10).to_string()}")
                logger.info("-------------------------------------------")

                # 1. 전체 셀 탐색
                for r_idx, row in df.iterrows():
                    for c_idx, cell_value in enumerate(row):
                        cell_str = str(cell_value).strip()

                        # 키워드 매칭 확인
                        if any(k in cell_str for k in KEYWORDS):
                            logger.info(
                                f"키워드 발견: '{cell_str}' at ({r_idx}, {c_idx})"
                            )

                            # Case A: 같은 열, 아래 행들에 데이터가 있는 경우 (테이블 형식 헤더)
                            # -> 이 경우는 보통 header=0으로 읽었을 때 처리되지만, header=None으로 읽었으므로
                            #    아래 행들을 검사해야 함. 하지만 폼 형식이 우선이므로 오른쪽 셀 먼저 확인.

                            # Case B: 바로 오른쪽 셀에 데이터가 있는 경우 (폼 형식)
                            if c_idx + 1 < len(df.columns):
                                target_val = df.iloc[r_idx, c_idx + 1]
                                if pd.notna(target_val):
                                    # 줄바꿈으로 구분된 여러 이름 처리
                                    names = (
                                        str(target_val)
                                        .replace("\r\n", "\n")
                                        .split("\n")
                                    )
                                    attendance_list = [
                                        n.strip() for n in names if n.strip()
                                    ]

                                    if attendance_list:
                                        logger.info(
                                            f"폼 형식 데이터 추출 완료: {len(attendance_list)}명"
                                        )
                                        all_attendees.extend(attendance_list)
                                        continue  # 다음 키워드 검색

                            # Case C: 테이블 형식 (아래 행들)
                            # 간단히 해당 열의 아래 모든 데이터를 긁어봄
                            # (헤더 바로 아래부터 끝까지)
                            col_data = (
                                df.iloc[r_idx + 1 :, c_idx]
                                .dropna()
                                .astype(str)
                                .tolist()
                            )
                            attendance_list = [n.strip() for n in col_data if n.strip()]
                            if attendance_list:
                                logger.info(
                                    f"테이블 형식 데이터 추출 완료: {len(attendance_list)}명"
                                )
                                all_attendees.extend(attendance_list)

            if not all_attendees:
                logger.warning(
                    f"'{file_path}'에서 키워드 {KEYWORDS}를 찾을 수 없습니다."
                )

            # 필터링 로직 추가
            filtered_attendees = []
            EXCLUDED_KEYWORDS = [
                "기간",
                "참석명단",
                "진행방식",
                "스터디 목표",
                "활동내용",
                "활동사진",
                "활동계획",
                "Webex 신청여부",
                "비고",
            ]

            REAL_EXCLUDED_KEYWORDS = [
                "기간",
                "참석명단",
                "진행방식",
                "스터디 목표",
                "활동내용",
                "활동사진",
                "활동계획",
                "Webex",
                "신청여부",
                "비고",
                "순번",
                "이름",
                "등록일시",
                "온라인",
                "오프라인",
            ]

            for name in all_attendees:
                clean_name = name.strip()
                if not clean_name:
                    continue

                # 1. 제외 키워드가 포함된 경우 건너뜀
                if any(k in clean_name for k in REAL_EXCLUDED_KEYWORDS):
                    continue

                # 2. 너무 긴 텍스트는 이름이 아닐 확률 높음 (예: 10자 이상)
                if len(clean_name) > 10:
                    continue

                # 3. 숫자로만 구성된 경우 제외
                if clean_name.isdigit():
                    continue

                # 4. 날짜 형식 (YYYY-MM-DD) 제외
                import re

                if re.match(r"\d{4}-\d{2}-\d{2}", clean_name):
                    continue

                filtered_attendees.append(clean_name)

            return filtered_attendees

        except Exception as e:
            logger.error(f"Pandas로 참석 명단 추출 중 오류 발생: {e}")
            # Fallback to openpyxl if available
            if not HAS_OPENPYXL:
                return []

    if HAS_OPENPYXL:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True)

            all_attendees = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                for row in ws.iter_rows(values_only=True):
                    for idx, cell_value in enumerate(row):
                        cell_str = str(cell_value).strip() if cell_value else ""

                        if any(k in cell_str for k in KEYWORDS):
                            # 오른쪽 셀 확인
                            if idx + 1 < len(row):
                                target_val = row[idx + 1]
                                if target_val:
                                    names = (
                                        str(target_val)
                                        .replace("\r\n", "\n")
                                        .split("\n")
                                    )
                                    attendance_list = [
                                        n.strip() for n in names if n.strip()
                                    ]
                                    if attendance_list:
                                        all_attendees.extend(attendance_list)

            # 필터링 로직 추가 (Pandas와 동일)
            filtered_attendees = []
            REAL_EXCLUDED_KEYWORDS = [
                "기간",
                "참석명단",
                "진행방식",
                "스터디 목표",
                "활동내용",
                "활동사진",
                "활동계획",
                "Webex",
                "신청여부",
                "비고",
                "순번",
                "이름",
                "등록일시",
            ]

            for name in all_attendees:
                clean_name = name.strip()
                if not clean_name:
                    continue

                if any(k in clean_name for k in REAL_EXCLUDED_KEYWORDS):
                    continue

                if len(clean_name) > 10:
                    continue

                if clean_name.isdigit():
                    continue

                filtered_attendees.append(clean_name)

            return filtered_attendees
        except Exception as e:
            logger.error(f"Openpyxl로 참석 명단 추출 중 오류 발생: {e}")
            return []

    logger.error(
        "Pandas 또는 Openpyxl이 설치되어 있지 않아 참석 명단을 추출할 수 없습니다."
    )
    return []


def parse_attendance_string(text: str) -> Optional[dict]:
    """참석자 문자열을 파싱합니다.
    예: "구미 2반 권택민" -> {"campus": "구미", "class_name": "2반", "name": "권택민"}
    """
    import re

    text = text.strip()

    # 1. 공백으로 구분된 경우 (기존 로직과 유사하지만 반 형식 확인 추가)
    parts = text.split()
    if len(parts) >= 3:
        # 가장 마지막 요소가 이름, 그 앞이 반, 그 앞이 캠퍼스라고 가정
        name = parts[-1]
        class_name = parts[-2]
        campus = parts[-3]

        if "반" in class_name:
            return {"campus": campus, "class_name": class_name, "name": name}

    # 2. 정규식으로 추출 (공백이 없거나 불규칙한 경우)
    # 예: "구미6반 김대규", "구미 6반김대규", "구미6반김대규", "구미_6반_김대규"
    # 캠퍼스(2글자 이상), 반(숫자+반), 이름(2글자 이상)
    match = re.search(r"([가-힣]{2,})[\s_]*(\d+반)[\s_]*([가-힣]{2,})", text)
    if match:
        return {
            "campus": match.group(1),
            "class_name": match.group(2),
            "name": match.group(3),
        }

    return None
